import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
c=tuple(sheet.columns)
#for c in sheet.columns:
for  cellobjec in c[3]:
	print(cellobjec.value)
	#sheet.append(tuple(cellobjec.value))
wb.create_sheet(index=1,title='cyy')
newsheet = wb.get_sheet_by_name('cyy')
for  cellobjec in c[3]:
	print(cellobjec.value)
newsheet.append(['y1','y2','y3','y4'])
for row in range(2,sheet.max_row+1):
	code=sheet['D'+str(row)].value
	print('code = '+str(code))
	print(list(str(code)))
	sheet.append(list(str(code)))
	sheet['E'+str(row)]=list(str(code))[0]
	sheet['F'+str(row)]=list(str(code))[1]
	sheet['G'+str(row)]=list(str(code))[2]
	sheet['H'+str(row)]=list(str(code))[3]

wb.save("samplesa.xlsx")
#print(type(cellobjec.value))
#for columns in sheet.iter_cols(min_col=4, max_col=4, max_row=sheet.max_row):
#   for cell in columns:
#       print(cell.value)